import openpyxl


class excel():
    @staticmethod
    def getDataFromExcel(rowNumber,colNumber):

        file = openpyxl.load_workbook("C:\\Users\\chand\\PycharmProjects\\OpenMRS_PythonProject\\Utilities\\OpenMRS_DataFile.xlsx")
        fileActive=file.active
        data=fileActive.cell(row=rowNumber,column=colNumber).value
        print(data)
        return data

    @staticmethod
    def setDataToExcel(rowNumber, colNumber, data):
        file = openpyxl.load_workbook(
            "C:\\Users\\chand\\PycharmProjects\\OpenMRS_PythonProject\\Utilities\\SuccessfulPatient_OpenMRS_DataFile.xlsx")
        sheet=file.get_sheet_by_name('Sheet1')
        cell=sheet.cell(row=rowNumber,column=colNumber)
        cell.value=data
        file.save("C:\\Users\\chand\\PycharmProjects\\OpenMRS_PythonProject\\Utilities\\SuccessfulPatient_OpenMRS_DataFile.xlsx")



        fileActive = file.active
        data = fileActive.cell(row=rowNumber, column=colNumber).value
        print(data)
        return data